"""Convert ACRIN 6664 XLSX metadata files to clean CSVs.

Reads the three Excel files provided by TCIA (no-polyp, 6-9mm, >=10mm),
normalizes column names, and writes a single combined CSV plus per-category
CSVs into the ``csv_metadata/`` directory.

The XLSX column layout for lesion files follows a repeating pattern of 5
sub-columns per lesion: location, size (mm), morphology, histology, and a
per-lesion flag.  This script extracts the largest polyp size per patient
for the combined output.

Usage:
    python scripts/convert_metadata.py \
        --input-dir raw_metadata/ACRIN_6664 \
        --output-dir csv_metadata
"""

import argparse
import csv
import os

import openpyxl


def read_no_polyp(xlsx_path):
    """Parse the no-polyp-found file.  Single column: TCIA Patient ID."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    patients = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        pid = row[0]
        if pid:
            patients.append({"patient_id": str(pid).strip(), "max_polyp_mm": 0,
                             "category": "no_polyp"})
    wb.close()
    return patients


def read_lesion_file(xlsx_path, category):
    """Parse a lesion file (6-9mm or >=10mm).

    Each row is a patient.  Lesion sub-columns repeat in groups of 5:
    [location, size_mm, morphology, histology, flag].  We extract the
    max polyp size across all lesions for each patient.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    patients = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        pid = row[0]
        if not pid:
            continue

        # Lesion data starts at column index 3 (after TCIA#, slice supine, slice prone)
        # Each lesion has 5 sub-columns; the size is at offset 1 within each group
        max_size = 0
        lesion_data = list(row[3:])
        for j in range(0, len(lesion_data), 5):
            size_val = lesion_data[j + 1] if (j + 1) < len(lesion_data) else None
            if size_val is not None:
                try:
                    max_size = max(max_size, float(size_val))
                except (ValueError, TypeError):
                    pass

        patients.append({"patient_id": str(pid).strip(),
                         "max_polyp_mm": max_size,
                         "category": category})
    wb.close()
    return patients


def main():
    parser = argparse.ArgumentParser(description="Convert ACRIN XLSX metadata to CSV")
    parser.add_argument("--input-dir", default="raw_metadata/ACRIN_6664",
                        help="Directory containing XLSX files")
    parser.add_argument("--output-dir", default="csv_metadata",
                        help="Output directory for CSVs")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    # Parse each Excel file
    no_polyp = read_no_polyp(
        os.path.join(args.input_dir, "TCIA-CTC-no-polyp-found.xlsx"))
    medium = read_lesion_file(
        os.path.join(args.input_dir, "TCIA-CTC-6-to-9-mm-polyps.xlsx"), "medium_6_9mm")
    large = read_lesion_file(
        os.path.join(args.input_dir, "TCIA-CTC-large-10-mm-polyps.xlsx"), "large_10mm_plus")

    # Write per-category CSVs
    fieldnames = ["patient_id", "max_polyp_mm", "category"]

    for name, data in [("no_polyp", no_polyp), ("medium_6_9mm", medium),
                       ("large_10mm_plus", large)]:
        path = os.path.join(args.output_dir, f"{name}.csv")
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        print(f"Wrote {len(data)} rows to {path}")

    # Write combined CSV with all patients
    combined = no_polyp + medium + large
    combined_path = os.path.join(args.output_dir, "acrin_combined.csv")
    with open(combined_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(combined)
    print(f"Wrote {len(combined)} combined rows to {combined_path}")

    # Summary
    print(f"\nSummary: {len(no_polyp)} no-polyp, {len(medium)} medium (6-9mm), "
          f"{len(large)} large (>=10mm) — {len(combined)} total")


if __name__ == "__main__":
    main()
